# app.py
# =====================================================
# GST DAYBOOK → TALLY WEB APPLICATION (FREE CLOUD VERSION)
# ORIGINAL LOGIC 100% PRESERVED
# =====================================================

import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import pandas as pd
import re
import io
import json
import xml.etree.ElementTree as ET

st.set_page_config(page_title="GST → Tally Automation", layout="wide")
st.title("📊 GST DayBook → Tally Automation Tool")

# ==============================
# FILE UPLOAD
# ==============================
uploaded_file = st.file_uploader("Upload DayBook Excel", type=["xlsx"])

# ==============================
# HELPER FUNCTIONS (UNCHANGED)
# ==============================
def normalize(text):
    return re.sub(r'[^a-z]', '', str(text).lower())

def is_bold(cell):
    return cell.font and cell.font.bold

green_fill = PatternFill("solid", fgColor="90EE90")
red_fill = PatternFill("solid", fgColor="FF7F7F")

if uploaded_file:
    wb = load_workbook(uploaded_file)
    ws = wb.active

    # ==============================
    # FIND HEADER ROW
    # ==============================
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(), start=1):
        values = [str(c.value).strip() if c.value else "" for c in row]
        if "Particulars" in values:
            header_row_idx = i
            break

    headers_map = {
        cell.value: idx
        for idx, cell in enumerate(ws[header_row_idx])
        if cell.value
    }

    # ==============================
    # GST COLUMN DETECTION
    # ==============================
    cgst_cols, sgst_cols, igst_cols = [], [], []
    for col_name, index in headers_map.items():
        norm = normalize(col_name)
        if "cgst" in norm:
            cgst_cols.append(index)
        elif "sgst" in norm:
            sgst_cols.append(index)
        elif "igst" in norm:
            igst_cols.append(index)

    # ==============================
    # CREATE CLEANED WORKBOOK (IN MEMORY)
    # ==============================
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Cleaned_Data"

    out_ws.append([
        "Date","Party Name","Stock Item","Voucher Type","Voucher No",
        "Voucher Ref No","GSTIN/UIN","Narration",
        "Quantity","Rate","UQC",
        "Taxable Value","CGST","SGST","IGST","Total Tax",
        "Round Off","Tax Rate","Invoice Value",
        "Result- Taxable Value","Result- Invoice Value"
    ])

    expected_taxable_total = calculated_taxable_total = 0
    expected_gross_total = calculated_invoice_total = 0

    current_total_cgst = current_total_sgst = current_total_igst = 0
    current_round_off = 0
    party_row_indexes = []

    for row in ws.iter_rows(min_row=header_row_idx + 1):
        particulars = row[headers_map["Particulars"]]
        qty_cell = row[headers_map["Quantity"]]
        rate_cell = row[headers_map["Rate"]]
        value_cell = row[headers_map["Value"]]
        gross_total_cell = row[headers_map["Gross Total"]]
        round_off_cell = row[headers_map["Round Off"]]

        if is_bold(particulars):
            if party_row_indexes:
                taxable_result = "Matched" if round(calculated_taxable_total,2)==round(expected_taxable_total,2) else "Not Matched"
                invoice_result = "Matched" if round(calculated_invoice_total,2)==round(expected_gross_total,2) else "Not Matched"

                fill1 = green_fill if taxable_result=="Matched" else red_fill
                fill2 = green_fill if invoice_result=="Matched" else red_fill

                for r in party_row_indexes:
                    out_ws.cell(row=r, column=20).value = taxable_result
                    out_ws.cell(row=r, column=20).fill = fill1
                    out_ws.cell(row=r, column=21).value = invoice_result
                    out_ws.cell(row=r, column=21).fill = fill2

            party_row_indexes = []
            calculated_taxable_total = calculated_invoice_total = 0

            expected_taxable_total = float(value_cell.value or 0)
            expected_gross_total = float(gross_total_cell.value or 0)

            current_total_cgst = sum(float(row[i].value or 0) for i in cgst_cols)
            current_total_sgst = sum(float(row[i].value or 0) for i in sgst_cols)
            current_total_igst = sum(float(row[i].value or 0) for i in igst_cols)
            current_round_off = float(round_off_cell.value or 0)

            raw_date = row[headers_map["Date"]].value
            current_date = raw_date.date() if isinstance(raw_date, datetime) else raw_date
            current_party = particulars.value
            current_voucher_type = row[headers_map["Voucher Type"]].value
            current_voucher_no = row[headers_map["Voucher No."]].value
            current_voucher_ref = row[headers_map["Voucher Ref. No."]].value
            current_gstin = row[headers_map["GSTIN/UIN"]].value
            current_narration = row[headers_map["Narration"]].value
            continue

        if is_bold(qty_cell):
            continue

        if rate_cell.value and "/" in str(rate_cell.value):
            rate_part, uqc = str(rate_cell.value).split("/",1)
            rate = float(rate_part)
        else:
            rate = float(rate_cell.value or 0)
            uqc = None

        qty = float(qty_cell.value or 0)
        taxable_value = round(qty * rate, 2)
        calculated_taxable_total += taxable_value

        ratio = taxable_value/expected_taxable_total if expected_taxable_total else 0
        cgst = round(ratio*current_total_cgst,2)
        sgst = round(ratio*current_total_sgst,2)
        igst = round(ratio*current_total_igst,2)
        round_off = round(ratio*current_round_off,2)

        total_tax = cgst+sgst+igst
        tax_rate = round((total_tax/taxable_value)*100,2) if taxable_value else 0
        invoice_value = round(taxable_value + total_tax + round_off,2)
        calculated_invoice_total += invoice_value

        out_ws.append([
            current_date,current_party,particulars.value,current_voucher_type,
            current_voucher_no,current_voucher_ref,current_gstin,current_narration,
            qty,rate,uqc,taxable_value,cgst,sgst,igst,total_tax,
            round_off,tax_rate,invoice_value,"",""])

        party_row_indexes.append(out_ws.max_row)

    cleaned_buffer = io.BytesIO()
    out_wb.save(cleaned_buffer)
    cleaned_buffer.seek(0)

    st.success("✅ Cleaned File Generated")
    st.download_button("⬇️ Download DayBook_Cleaned.xlsx", cleaned_buffer, "DayBook_Cleaned.xlsx")

    # =====================================================
    # CREATE FILE_TO_TALLY (SAME ORIGINAL GST ENGINE LOGIC)
    # =====================================================

    cleaned_buffer.seek(0)
    cleaned_wb = load_workbook(cleaned_buffer)
    cleaned_ws = cleaned_wb.active

    tally_wb = Workbook()
    tally_ws = tally_wb.active
    tally_ws.title = "File_to_Tally"

    required_columns = [
        "Date","Voucher Type","Voucher No","Voucher Ref No","Party Name",
        "GSTIN/UIN","Stock Item","Quantity","Rate",
        "Taxable Value","CGST","SGST","IGST",
        "Tax Rate","Invoice Value"
    ]

    header_map = {
        cleaned_ws.cell(row=1, column=c).value: c
        for c in range(1, cleaned_ws.max_column + 1)
    }

    tally_ws.append(required_columns)

    # ==============================
    # USER-CONTROLLED DECREASE PERCENTAGE (CASE-WISE)
    # ==============================

    DECREASE_PERCENTAGE = st.number_input(
        "🔻 Decrease Percentage (case-wise)",
        min_value=0,
        max_value=100,
        value=20,
        step=1,
        help="Change this as per case: 0, 10, 20, 30, 40 etc."
    )

    factor = (100 - DECREASE_PERCENTAGE) / 100

    for row_num in range(2, cleaned_ws.max_row + 1):

        narration_value = cleaned_ws.cell(row=row_num, column=header_map["Narration"]).value
        narration_text = str(narration_value).lower() if narration_value else ""
        narration_text_clean = narration_text.replace(" ", "")

        original_party = cleaned_ws.cell(row=row_num, column=header_map["Party Name"]).value
        original_gstin = cleaned_ws.cell(row=row_num, column=header_map["GSTIN/UIN"]).value
        tax_rate = float(cleaned_ws.cell(row=row_num, column=header_map["Tax Rate"]).value or 0)

        if "gst" not in narration_text_clean:
            party_value = "Cash"
            gstin_value = None
        else:
            party_value = original_party
            gstin_value = original_gstin

        qty_value = float(cleaned_ws.cell(row=row_num, column=header_map["Quantity"]).value or 0)
        rate_value = float(cleaned_ws.cell(row=row_num, column=header_map["Rate"]).value or 0)

        if not gstin_value:
            qty_value = round(qty_value * factor, 2)
            rate_value = round(rate_value * factor, 2)

        taxable_value = round(qty_value * rate_value, 2)

        cgst = sgst = igst = 0

        if not gstin_value:
            cgst = round((taxable_value * (tax_rate/100)) / 2, 2)
            sgst = round((taxable_value * (tax_rate/100)) / 2, 2)
        else:
            gstin_str = str(gstin_value).strip()
            if gstin_str[:2] == "33":
                cgst = round((taxable_value * (tax_rate/100)) / 2, 2)
                sgst = round((taxable_value * (tax_rate/100)) / 2, 2)
            else:
                igst = round(taxable_value * (tax_rate/100), 2)

        invoice_value = round(taxable_value + cgst + sgst + igst, 2)

        new_row = []

        for col_name in required_columns:
            if col_name == "Party Name":
                new_row.append(party_value)
            elif col_name == "GSTIN/UIN":
                new_row.append(gstin_value)
            elif col_name == "Quantity":
                new_row.append(qty_value)
            elif col_name == "Rate":
                new_row.append(rate_value)
            elif col_name == "Taxable Value":
                new_row.append(taxable_value)
            elif col_name == "CGST":
                new_row.append(cgst)
            elif col_name == "SGST":
                new_row.append(sgst)
            elif col_name == "IGST":
                new_row.append(igst)
            elif col_name == "Tax Rate":
                new_row.append(tax_rate)
            elif col_name == "Invoice Value":
                new_row.append(invoice_value)
            else:
                col_index = header_map.get(col_name)
                value = cleaned_ws.cell(row=row_num, column=col_index).value if col_index else None
                if col_name == "Date" and isinstance(value, datetime):
                    value = value.date()
                new_row.append(value)

        tally_ws.append(new_row)

    tally_buffer = io.BytesIO()
    tally_wb.save(tally_buffer)
    tally_buffer.seek(0)

    st.success("✅ File_to_Tally Generated")
    st.download_button("⬇️ Download File_to_Tally.xlsx", tally_buffer, "File_to_Tally.xlsx")

    # =====================================================
    # PERFORMANCE OPTIMIZATION + DASHBOARD PREVIEW
    # =====================================================

    # Convert cleaned sheet to pandas (FASTER aggregation for 5,000+ rows)
    cleaned_buffer.seek(0)
    df = pd.read_excel(cleaned_buffer)

    # Ensure Date column is datetime
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")

    # Create Month column
    df["Month"] = df["Date"].dt.to_period("M").astype(str)

    # ==============================
    # DASHBOARD PREVIEW (CLEANED + FILE_TO_TALLY)
    # ==============================

    st.markdown("---")
    st.subheader("📊 Monthly GST Dashboard Preview")

    col1, col2 = st.columns(2)

    # -------- CLEANED DATA DASHBOARD --------
    with col1:
        st.markdown("### 🧾 DayBook Cleaned Summary")

        cleaned_buffer.seek(0)
        df_cleaned = pd.read_excel(cleaned_buffer)
        df_cleaned["Date"] = pd.to_datetime(df_cleaned["Date"], errors="coerce")
        df_cleaned["Month"] = df_cleaned["Date"].dt.to_period("M").astype(str)

        cleaned_summary = (
            df_cleaned.groupby("Month")[["CGST", "SGST", "IGST", "Taxable Value", "Invoice Value"]]
            .sum()
            .reset_index()
        )

        st.dataframe(cleaned_summary, use_container_width=True)
        st.line_chart(cleaned_summary.set_index("Month")["Invoice Value"])

    # -------- FILE TO TALLY DASHBOARD --------
    with col2:
        st.markdown("### 📤 File to Tally Summary")

        tally_buffer.seek(0)
        df_tally = pd.read_excel(tally_buffer)
        df_tally["Date"] = pd.to_datetime(df_tally["Date"], errors="coerce")
        df_tally["Month"] = df_tally["Date"].dt.to_period("M").astype(str)

        tally_summary = (
            df_tally.groupby("Month")[["CGST", "SGST", "IGST", "Taxable Value", "Invoice Value"]]
            .sum()
            .reset_index()
        )

        st.dataframe(tally_summary, use_container_width=True)
        st.line_chart(tally_summary.set_index("Month")["Invoice Value"])

    # DOWNLOAD COMBINED DASHBOARD
    dashboard_buffer = io.BytesIO()
    with pd.ExcelWriter(dashboard_buffer, engine="openpyxl") as writer:
        cleaned_summary.to_excel(writer, sheet_name="DayBook_Cleaned", index=False)
        tally_summary.to_excel(writer, sheet_name="File_to_Tally", index=False)

    dashboard_buffer.seek(0)

    st.download_button(
        "⬇️ Download Monthly_GST_Dashboard.xlsx",
        dashboard_buffer,
        "Monthly_GST_Dashboard.xlsx"
    )

    st.info("⚡ Optimized dashboards generated separately for Cleaned data and File-to-Tally data (large files supported).")
